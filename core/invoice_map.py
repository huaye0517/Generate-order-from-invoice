# -*- coding: utf-8 -*-
"""开票信息映射"""
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple
import openpyxl


@dataclass
class InvoiceRecord:
    """开票信息一行完整记录"""
    contract_name: str
    company_name: str
    address: str
    signatory: str
    phone: str
    bank: str
    account: str
    credit_code: str


def _cell_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value).strip()


def load_invoice_mapping(template_path: str) -> List[Tuple[str, str]]:
    """读取开票信息表 A列(合同名称) -> C列(单位名称)"""
    records = load_invoice_records(template_path)
    return [(r.contract_name, r.company_name) for r in records if r.contract_name or r.company_name]


def load_invoice_records(template_path: str) -> List[InvoiceRecord]:
    """读取开票信息全部记录"""
    wb = openpyxl.load_workbook(template_path, read_only=True, data_only=True)
    if "开票信息" not in wb.sheetnames:
        wb.close()
        return []
    ws = wb["开票信息"]
    records: List[InvoiceRecord] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        contract = _cell_str(row[0]).lstrip("\ufeff") if len(row) > 0 else ""
        company = _cell_str(row[2]) if len(row) > 2 else ""
        if not contract and not company:
            continue
        records.append(
            InvoiceRecord(
                contract_name=contract,
                company_name=company,
                address=_cell_str(row[3]) if len(row) > 3 else "",
                signatory=_cell_str(row[4]) if len(row) > 4 else "",
                phone=_cell_str(row[5]) if len(row) > 5 else "",
                bank=_cell_str(row[6]) if len(row) > 6 else "",
                account=_cell_str(row[7]) if len(row) > 7 else "",
                credit_code=_cell_str(row[8]) if len(row) > 8 else "",
            )
        )
    wb.close()
    return records


def resolve_company_name(customer_name: str, mapping: List[Tuple[str, str]]) -> str:
    """目录客户名 -> 公司名（开票信息 C 列）"""
    cat = str(customer_name).strip()
    if not cat:
        return cat

    for a_str, c_str in mapping:
        if a_str and a_str.lower() == cat.lower() and c_str:
            return c_str

    best: Optional[str] = None
    best_len = -1
    for a_str, c_str in mapping:
        if not a_str or not c_str:
            continue
        a_lower = a_str.lower()
        cat_lower = cat.lower()
        if cat_lower.startswith(a_lower) or a_lower in cat_lower:
            if len(a_str) > best_len:
                best = c_str
                best_len = len(a_str)
    if best:
        return best

    return cat


def lookup_invoice_record(
    company_name: str,
    customer_names: List[str],
    records: List[InvoiceRecord],
) -> Optional[InvoiceRecord]:
    """
    按单位名称或客户名查找开票信息。
    优先 C 列精确匹配，其次 A 列精确/前缀匹配。
    """
    keys = [str(company_name).strip()] + [str(n).strip() for n in customer_names if n]
    keys = [k for k in keys if k]

    # C 列精确匹配
    for key in keys:
        key_lower = key.lower()
        for rec in records:
            if rec.company_name and rec.company_name.lower() == key_lower:
                return rec

    # A 列精确匹配
    for key in keys:
        key_lower = key.lower()
        for rec in records:
            if rec.contract_name and rec.contract_name.lower() == key_lower:
                return rec

    # A 列前缀/包含匹配
    best: Optional[InvoiceRecord] = None
    best_len = -1
    for key in keys:
        key_lower = key.lower()
        for rec in records:
            a = rec.contract_name
            if not a:
                continue
            a_lower = a.lower()
            if key_lower.startswith(a_lower) or a_lower in key_lower:
                if len(a) > best_len:
                    best = rec
                    best_len = len(a)
    return best


def buyer_info_from_record(record: Optional[InvoiceRecord], e2_value: str) -> Dict[str, str]:
    """生成需方信息（语义键，由 template_trim 按 layout 写入对应单元格）"""
    company = e2_value
    if record and record.company_name:
        company = record.company_name

    if record:
        return {
            "party": company,
            "company": company,
            "address": record.address,
            "signatory": record.signatory,
            "phone": record.phone,
            "bank": record.bank,
            "account": record.account,
            "credit_code": record.credit_code,
        }
    return {
        "party": company,
        "company": company,
        "address": "",
        "signatory": "",
        "phone": "",
        "bank": "",
        "account": "",
        "credit_code": "",
    }


def buyer_cells_from_record(record: Optional[InvoiceRecord], e2_value: str) -> Dict[str, str]:
    """兼容旧接口，返回 buyer_info_from_record 结果"""
    return buyer_info_from_record(record, e2_value)
