# -*- coding: utf-8 -*-
"""自动分析销售订单合同模板布局，供裁剪与写入适配使用"""
import re
from typing import Dict, List, Optional, Tuple

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .template_layout import TemplateLayout

# 产品行 B 列典型公式特征
PRODUCT_ROW_PATTERN = re.compile(r'=IF\s*\(\s*ROW\s*\(\s*A\d+\s*\)', re.I)
SUM_PATTERN = re.compile(
    r"=SUM\s*\(\s*([A-Z]+)(\d+)\s*:\s*([A-Z]+)(\d+)\s*\)",
    re.I,
)
# 大写金额公式特征
UPPERCASE_FORMULA_PATTERN = re.compile(r"DOLLAR|dbnum2|通用格式", re.I)

# 需方字段标签 -> 语义键（避免单字匹配导致误识别）
BUYER_LABEL_KEYS: List[Tuple[str, Tuple[str, ...]]] = [
    ("company", (
        "单位名称", "公司名称", "甲方单位", "甲方名称", "企业名称",
    )),
    ("address", (
        "单位地址", "公司地址", "通讯地址", "注册地址", "企业地址",
    )),
    ("signatory", (
        "代表签字", "法定代表", "代表人", "授权代表", "甲方代表", "乙方代表",
    )),
    ("phone", (
        "电话", "电    话", "电  话", "联系电话", "手机", "传真",
    )),
    ("bank", (
        "开户银行", "开户行",
    )),
    ("account", (
        "账号", "账    号", "账  号", "银行账号", "银行账户", "帐号", "帐    号",
    )),
    ("credit_code", (
        "统一社会信用代码", "社会信用代码", "信用代码", "纳税人识别号", "税务登记",
    )),
]

PARTY_LABELS = ("需求方", "甲  方", "甲方", "需方", "买方")
_BUYER_HEADER_PREFIXES = ("需方", "需求方", "甲方", "买方")
_SUPPLIER_HEADER_PREFIXES = ("供方", "供货方", "卖方", "乙方")


def _cell_str(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_label(text: str) -> str:
    """标签归一化：去空格、全角冒号"""
    return text.replace(" ", "").replace("：", "").replace(":", "")


def _row_is_empty(ws: Worksheet, row: int) -> bool:
    for col in range(1, ws.max_column + 1):
        if ws.cell(row, col).value not in (None, ""):
            return False
    return True


def _find_template_sheet(wb) -> str:
    if "模板" in wb.sheetnames:
        return "模板"
    for name in wb.sheetnames:
        ws = wb[name]
        if _find_total_row(ws) is not None:
            return name
    raise ValueError("未在模板文件中识别到合同工作表（需包含「合计」行及产品公式区）")


def _find_party_cell(ws: Worksheet) -> str:
    for row in range(1, min(15, ws.max_row + 1)):
        for col in range(1, min(8, ws.max_column + 1)):
            label = _cell_str(ws.cell(row, col).value)
            if not label:
                continue
            if any(p in label.replace(" ", "") for p in PARTY_LABELS):
                value_col = col + 2
                if value_col <= ws.max_column:
                    letter = get_column_letter(value_col)
                    return f"{letter}{row}"
    return "E2"


def _find_total_row(ws: Worksheet) -> Optional[int]:
    for row in range(1, ws.max_row + 1):
        for col in range(1, min(8, ws.max_column + 1)):
            text = _normalize_label(_cell_str(ws.cell(row, col).value))
            if text == "合计" or (text.endswith("合计") and "大写" not in text and "总合计" not in text):
                return row
    return None


def _parse_sum_range(ws: Worksheet, total_row: int) -> Optional[Tuple[int, int, int, int]]:
    """从合计行解析 SUM 公式，返回 (qty_col, amt_col, first_row, last_row)"""
    qty_col = None
    amt_col = None
    first_row = None
    last_row = None

    for col in range(1, ws.max_column + 1):
        val = ws.cell(total_row, col).value
        if not val or not isinstance(val, str) or not val.upper().startswith("=SUM"):
            continue
        m = SUM_PATTERN.match(val.strip())
        if not m:
            continue
        _, r1, _, r2 = m.groups()
        row_start = int(r1)
        row_end = int(r2)
        if first_row is None:
            first_row, last_row = row_start, row_end
        if qty_col is None:
            qty_col = col
        else:
            amt_col = col

    if first_row is None or last_row is None:
        return None
    if qty_col is None:
        qty_col = 7
    if amt_col is None:
        amt_col = 9
    return qty_col, amt_col, first_row, last_row


def _is_product_formula_row(ws: Worksheet, row: int) -> bool:
    for col in range(1, min(6, ws.max_column + 1)):
        val = ws.cell(row, col).value
        if val and isinstance(val, str) and PRODUCT_ROW_PATTERN.search(val):
            return True
    return False


def _refine_product_range(ws: Worksheet, first_row: int, last_row: int) -> Tuple[int, int]:
    while first_row <= last_row and not _is_product_formula_row(ws, first_row):
        first_row += 1
    while last_row >= first_row and not _is_product_formula_row(ws, last_row):
        last_row -= 1
    if first_row > last_row:
        raise ValueError("未能识别产品明细公式区")
    return first_row, last_row


def _find_grand_total_row(ws: Worksheet, total_row: int) -> int:
    for row in range(total_row + 1, min(total_row + 5, ws.max_row + 1)):
        for col in range(1, min(8, ws.max_column + 1)):
            text = _cell_str(ws.cell(row, col).value)
            if "总合计" in text or "大写" in text:
                return row
    return total_row + 1


def _find_grand_formula_info(ws: Worksheet, grand_row: int) -> Tuple[int, int, int]:
    """
    识别总合计行的大写金额公式列及其合并范围。
    返回 (formula_col, merge_min_col, merge_max_col)。
    """
    formula_col = 5
    merge_min, merge_max = 5, 5

    for col in range(1, ws.max_column + 1):
        val = ws.cell(grand_row, col).value
        if isinstance(val, str) and val.startswith("=") and UPPERCASE_FORMULA_PATTERN.search(val):
            formula_col = col
            break

    for m in ws.merged_cells.ranges:
        if m.min_row == grand_row and m.min_col <= formula_col <= m.max_col:
            merge_min = m.min_col
            merge_max = m.max_col
            break

    return formula_col, merge_min, merge_max


def _detect_product_row_merges(ws: Worksheet, product_first_row: int) -> List[Tuple[int, int]]:
    """从首行产品行识别单行跨列合并（删行后易产生幽灵合并）"""
    merges: List[Tuple[int, int]] = []
    for m in ws.merged_cells.ranges:
        if m.min_row == product_first_row and m.max_row == product_first_row:
            merges.append((m.min_col, m.max_col))
    return merges


def _find_supplier_start_col(ws: Worksheet, header_row: int) -> int:
    """识别供方区块起始列"""
    for col in range(1, min(12, ws.max_column + 1)):
        text = _normalize_label(_cell_str(ws.cell(header_row, col).value))
        if any(text.startswith(_normalize_label(p)) for p in _SUPPLIER_HEADER_PREFIXES):
            return col
    return 6


def _find_buyer_header_row(ws: Worksheet, start_row: int) -> Optional[int]:
    for row in range(start_row, ws.max_row + 1):
        for col in range(1, min(10, ws.max_column + 1)):
            text = _normalize_label(_cell_str(ws.cell(row, col).value))
            if any(text.startswith(_normalize_label(p)) for p in _BUYER_HEADER_PREFIXES):
                return row
    return None


def _detect_buyer_label_col(ws: Worksheet, start_row: int, supplier_col: int) -> int:
    """在需方区块内识别标签列（含「单位名称」等字段的行）"""
    for row in range(start_row, min(start_row + 12, ws.max_row + 1)):
        for col in range(1, supplier_col):
            text = _normalize_label(_cell_str(ws.cell(row, col).value))
            for _key, patterns in BUYER_LABEL_KEYS:
                if any(text.startswith(_normalize_label(p)) for p in patterns):
                    return col
    return 3


def _resolve_value_column(ws: Worksheet, row: int, col: int) -> int:
    for merged in ws.merged_cells.ranges:
        if (
            merged.min_row <= row <= merged.max_row
            and merged.min_col <= col <= merged.max_col
        ):
            return merged.min_col
    return col


def _is_in_label_merge(ws: Worksheet, row: int, col: int, label_col: int) -> bool:
    """判断单元格是否属于从标签列起始的合并区域"""
    for merged in ws.merged_cells.ranges:
        if (
            merged.min_row <= row <= merged.max_row
            and merged.min_col == label_col
            and merged.min_col <= col <= merged.max_col
        ):
            return True
    return False


def _label_matches(label: str, patterns: Tuple[str, ...]) -> bool:
    norm = _normalize_label(label)
    return any(norm.startswith(_normalize_label(p)) for p in patterns)


def _find_buyer_fields(
    ws: Worksheet,
    start_row: int,
    label_col: int,
    supplier_col: int,
) -> Tuple[Dict[str, int], int]:
    """扫描需方字段行，返回 {语义键: 行号} 及值写入列"""
    fields: Dict[str, int] = {}
    value_col = label_col + 2

    for row in range(start_row, ws.max_row + 1):
        label = _cell_str(ws.cell(row, label_col).value)
        if not label:
            continue
        label_norm = _normalize_label(label)
        if any(label_norm.startswith(_normalize_label(p)) for p in _SUPPLIER_HEADER_PREFIXES):
            break
        hit_supplier = False
        for col in range(supplier_col, min(supplier_col + 2, ws.max_column + 1)):
            side_text = _normalize_label(_cell_str(ws.cell(row, col).value))
            if any(side_text.startswith(_normalize_label(p)) for p in _SUPPLIER_HEADER_PREFIXES):
                hit_supplier = True
                break
        if hit_supplier:
            break

        for key, patterns in BUYER_LABEL_KEYS:
            if key in fields:
                continue
            if _label_matches(label, patterns):
                fields[key] = row
                formula_col = None
                fallback_col = None
                for col in range(label_col + 1, supplier_col):
                    cell = ws.cell(row, col)
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        if formula_col is None:
                            formula_col = _resolve_value_column(ws, row, col)
                    elif cell.value and not _cell_str(cell.value).startswith("="):
                        if fallback_col is None and not _is_in_label_merge(ws, row, col, label_col):
                            fallback_col = _resolve_value_column(ws, row, col)
                if formula_col is not None:
                    value_col = formula_col
                elif fallback_col is not None:
                    value_col = fallback_col
                break

    return fields, value_col


def _find_terms_last_row(ws: Worksheet, grand_row: int, buyer_header: Optional[int]) -> int:
    """识别条款区最后一行（含内容的行）"""
    end_row = grand_row
    scan_until = buyer_header if buyer_header else ws.max_row
    for row in range(grand_row + 1, scan_until):
        for col in range(1, min(8, ws.max_column + 1)):
            text = _cell_str(ws.cell(row, col).value)
            if text:
                end_row = row
                break
    return end_row


def _find_empty_rows_before_buyer(ws: Worksheet, from_row: int, buyer_header_row: int) -> List[int]:
    empty_rows: List[int] = []
    for row in range(from_row, buyer_header_row):
        if _row_is_empty(ws, row):
            empty_rows.append(row)
    return empty_rows


def analyze_template_layout(template_path: str, sheet_name: Optional[str] = None) -> TemplateLayout:
    """
    分析上传的销售订单合同模板，自动识别产品区、合计区、需方区位置。
    不同模板只要结构类似（产品公式区 + 合计 + 总合计大写 + 需方信息表），即可自动适配。
    """
    wb = openpyxl.load_workbook(template_path, data_only=False)
    try:
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表不存在: {sheet_name}")
            ws = wb[sheet_name]
            used_sheet = sheet_name
        else:
            used_sheet = _find_template_sheet(wb)
            ws = wb[used_sheet]

        total_row = _find_total_row(ws)
        if total_row is None:
            raise ValueError("未找到「合计」行，请确认模板格式")

        parsed = _parse_sum_range(ws, total_row)
        if parsed is None:
            raise ValueError("合计行未找到 SUM 公式，无法识别产品区范围")

        qty_col, amt_col, first_row, last_row = parsed
        first_row, last_row = _refine_product_range(ws, first_row, last_row)

        party_cell = _find_party_cell(ws)
        grand_total_row = _find_grand_total_row(ws, total_row)
        grand_formula_col, grand_merge_min, grand_merge_max = _find_grand_formula_info(
            ws, grand_total_row
        )
        product_row_merges = _detect_product_row_merges(ws, first_row)

        buyer_header = _find_buyer_header_row(ws, grand_total_row + 1)
        supplier_col = _find_supplier_start_col(ws, buyer_header) if buyer_header else 6
        terms_last = _find_terms_last_row(ws, grand_total_row, buyer_header)

        empty_before_buyer: List[int] = []
        buyer_fields: Dict[str, int] = {}
        buyer_label_col = 3
        buyer_value_col = 5

        if buyer_header:
            buyer_label_col = _detect_buyer_label_col(ws, buyer_header + 1, supplier_col)
            empty_before_buyer = _find_empty_rows_before_buyer(ws, terms_last + 1, buyer_header)
            buyer_fields, buyer_value_col = _find_buyer_fields(
                ws, buyer_header + 1, buyer_label_col, supplier_col
            )
        else:
            buyer_fields, buyer_value_col = _find_buyer_fields(ws, grand_total_row + 1, 3, 6)

        if not buyer_fields:
            raise ValueError("未找到需方信息区（单位名称/地址等标签）")

        return TemplateLayout(
            template_sheet=used_sheet,
            party_cell=party_cell,
            product_first_row=first_row,
            product_last_row=last_row,
            total_row=total_row,
            qty_sum_col=qty_col,
            amt_sum_col=amt_col,
            grand_total_row=grand_total_row,
            buyer_value_col=buyer_value_col,
            buyer_fields=buyer_fields,
            empty_rows_before_buyer=empty_before_buyer,
            grand_formula_col=grand_formula_col,
            grand_formula_merge_min_col=grand_merge_min,
            grand_formula_merge_max_col=grand_merge_max,
            product_row_merges=product_row_merges,
            buyer_label_col=buyer_label_col,
            buyer_header_row=buyer_header or 0,
            supplier_start_col=supplier_col,
            terms_last_row=terms_last,
        )
    finally:
        wb.close()
