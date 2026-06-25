# -*- coding: utf-8 -*-
"""模板页自适应裁剪：按分析结果删除留白行，保留表格线条"""
from typing import Dict, Optional, Tuple

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .template_layout import TemplateLayout

BUYER_INFO_KEYS = (
    "company",
    "address",
    "signatory",
    "phone",
    "bank",
    "account",
    "credit_code",
)


def _row_is_empty(ws, row: int) -> bool:
    for col in range(1, ws.max_column + 1):
        if ws.cell(row, col).value not in (None, ""):
            return False
    return True


def _replace_row_in_formula(formula: str, old_row: int, new_row: int) -> str:
    if not formula or not isinstance(formula, str):
        return formula
    text = formula.replace(f"I{old_row}", f"I{new_row}")
    text = text.replace(f"G{old_row}", f"G{new_row}")
    return text


def _unmerge_covering(ws: Worksheet, row: int, col: int) -> None:
    """解除覆盖指定单元格的合并，以便写入值"""
    to_remove = [
        m for m in list(ws.merged_cells.ranges)
        if m.min_row <= row <= m.max_row and m.min_col <= col <= m.max_col
    ]
    for m in to_remove:
        try:
            ws.unmerge_cells(str(m))
        except (KeyError, TypeError):
            _safe_remove_merge(ws, m)


def _write_buyer_value(ws: Worksheet, row: int, col: int, value: str) -> None:
    """写入需方字段值（处理删行后合并单元格错位）"""
    _unmerge_covering(ws, row, col)
    ws.cell(row, col).value = value


def _save_grand_e_info(ws: Worksheet, grand_total_row: int) -> Tuple[Optional[str], int, int]:
    """
    删行前保存总合计行 E 列的大写公式及其合并范围。
    返回 (formula, merge_min_col, merge_max_col)。
    """
    formula = ws.cell(grand_total_row, 5).value
    if not formula or not isinstance(formula, str) or not formula.startswith("="):
        formula = None

    merge_min_col, merge_max_col = 5, 5
    for m in ws.merged_cells.ranges:
        if m.min_row == grand_total_row and m.min_col == 5:
            merge_min_col = m.min_col
            merge_max_col = m.max_col
            break

    return formula, merge_min_col, merge_max_col


def _safe_remove_merge(ws: Worksheet, merge_range) -> None:
    """
    安全移除合并区域声明，不触碰单元格数据（避免幽灵合并引起的 KeyError）。
    ws.unmerge_cells 在单元格不存在时会崩溃，直接操作 merged_cells.ranges 更安全。
    """
    try:
        ws.merged_cells.ranges.discard(merge_range)
    except Exception:
        pass


def _purge_ghost_de_merges(ws: Worksheet, row_from: int, row_to: int) -> None:
    """
    清除 openpyxl delete_rows 遗留的幽灵 D:E 合并单元格。
    这些幽灵合并来自被删除产品行（每行独立的 D:E 合并），
    会残留在相同行号位置，遮盖总合计大写公式及需方信息格。
    仅清除 min_col=D(4)、max_col=E(5)、单行跨列的合并。
    """
    to_remove = [
        m for m in list(ws.merged_cells.ranges)
        if (
            row_from <= m.min_row <= row_to
            and m.min_col == 4
            and m.max_col == 5
            and m.min_row == m.max_row
        )
    ]
    for m in to_remove:
        _safe_remove_merge(ws, m)


def _restore_grand_e_formula(
    ws: Worksheet,
    grand_row: int,
    formula: str,
    old_total_row: int,
    new_total_row: int,
    merge_min_col: int,
    merge_max_col: int,
) -> None:
    """
    在总合计行恢复大写公式：
    1. 清除该行 merge_min_col～merge_max_col 范围内的所有合并（包括幽灵合并）
    2. 写入行号已修正的公式
    3. 若原来是多列合并，则恢复合并
    """
    to_remove = [
        m for m in list(ws.merged_cells.ranges)
        if (
            m.min_row == grand_row
            and m.min_col >= merge_min_col
            and m.max_col <= merge_max_col
        )
    ]
    for m in to_remove:
        _safe_remove_merge(ws, m)

    corrected = _replace_row_in_formula(formula, old_total_row, new_total_row)
    ws.cell(grand_row, merge_min_col).value = corrected

    if merge_max_col > merge_min_col:
        start_col = get_column_letter(merge_min_col)
        end_col = get_column_letter(merge_max_col)
        ws.merge_cells(f"{start_col}{grand_row}:{end_col}{grand_row}")


def trim_template_sheet(
    dest_path: str,
    layout: TemplateLayout,
    buyer_info: Dict[str, str],
    product_line_count: int,
) -> None:
    """
    根据自动分析的 layout 裁剪模板：
    1. 删除产品区多余空行
    2. 删除需方区块前的空白行
    3. 更新合计/总合计公式
    4. 写入需方静态信息（打开即显示，无需点击重算）
    """
    wb = openpyxl.load_workbook(dest_path)
    ws = wb[layout.template_sheet]

    last_data_row = layout.last_product_row(product_line_count)
    delete_count = layout.rows_to_delete_count(product_line_count)

    # 删行前保存总合计大写公式（E208 类），避免 openpyxl delete_rows 幽灵合并损坏
    e_grand_formula, e_merge_min, e_merge_max = _save_grand_e_info(
        ws, layout.grand_total_row
    )

    if delete_count > 0:
        ws.delete_rows(last_data_row + 1, delete_count)

    total_row = layout.total_row - delete_count
    qty_col = layout.qty_sum_col
    amt_col = layout.amt_sum_col
    qty_letter = get_column_letter(qty_col)
    amt_letter = get_column_letter(amt_col)

    ws.cell(total_row, qty_col).value = (
        f"=SUM({qty_letter}{layout.product_first_row}:{qty_letter}{last_data_row})"
    )
    ws.cell(total_row, amt_col).value = (
        f"=SUM({amt_letter}{layout.product_first_row}:{amt_letter}{last_data_row})"
    )

    grand_row = layout.grand_total_row - delete_count
    ws.cell(grand_row, amt_col).value = f"={amt_letter}{total_row}"

    # 清除 last_data_row+1 到 grand_row 之间的幽灵 D:E 合并（产品行删除遗留）
    if delete_count > 0:
        _purge_ghost_de_merges(ws, last_data_row + 1, grand_row)

    # 恢复总合计大写公式（修正行引用 + 重建合并）
    if e_grand_formula:
        _restore_grand_e_formula(
            ws, grand_row, e_grand_formula,
            layout.total_row, total_row,
            e_merge_min, e_merge_max,
        )

    gap_deleted = 0
    for gap_row in sorted(layout.empty_rows_before_buyer, reverse=True):
        shifted = gap_row - delete_count
        if shifted > 0 and shifted <= ws.max_row and _row_is_empty(ws, shifted):
            ws.delete_rows(shifted, 1)
            gap_deleted += 1

    row_shift = delete_count + gap_deleted
    buyer_rows = {
        key: row - row_shift
        for key, row in layout.buyer_fields.items()
    }

    # 清除需方区域的幽灵 D:E 合并，避免干扰标签列合并结构
    if buyer_rows:
        valid_rows = [r for r in buyer_rows.values() if r > 0]
        if valid_rows:
            _purge_ghost_de_merges(ws, min(valid_rows) - 1, max(valid_rows) + 1)

    party_cell = layout.party_cell
    party_row = ws[party_cell].row
    party_col = ws[party_cell].column
    _unmerge_covering(ws, party_row, party_col)
    ws.cell(party_row, party_col).value = buyer_info.get(
        "party", buyer_info.get("company", "")
    )

    value_col = layout.buyer_value_col
    for key in BUYER_INFO_KEYS:
        row = buyer_rows.get(key)
        if row and row <= ws.max_row:
            _write_buyer_value(ws, row, value_col, buyer_info.get(key, ""))

    # 重置工作表视图，打开时从第1行开始显示
    ws.sheet_view.topLeftCell = "A1"
    if ws.sheet_view.selection:
        ws.sheet_view.selection[0].activeCell = "A1"
        ws.sheet_view.selection[0].sqref = "A1"

    wb.save(dest_path)
    wb.close()
